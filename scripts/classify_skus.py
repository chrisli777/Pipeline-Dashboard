#!/usr/bin/env python3
"""
Phase 3A: ABC/XYZ SKU Classification Engine
============================================
Parses WMS Excel exports from Extensiv and computes:
  - ABC classification (annual consumption value, Pareto 80/96/100)
  - XYZ classification (coefficient of variation of monthly demand)
  - SKU master data (unit cost, weight, dimensions)

Outputs:
  - scripts/014_sku_classification_and_master_data.sql
  - Console summary of classification results

Data sources:
  - ItemActivityInventoryDepletion.xlsx  (Avg Ship/Wk per SKU)
  - ViewItem (3).xlsx                    (Cost, Weight, Dimensions per SKU)
  - Item_Activity_Report (1-6).xlsx      (Daily transaction detail for CV calc)

IMPORTANT: Excludes inventory adjustments (ref# containing 'adjust'/'adj')
           and warehouse moves (ref# starting with 'mv')
"""

import os
import sys
import math
import statistics
from datetime import datetime
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ─── Configuration ───────────────────────────────────────────────────────────

DOWNLOADS = Path(os.path.expanduser("~/Downloads"))
SCRIPT_DIR = Path(__file__).parent
OUTPUT_SQL = SCRIPT_DIR / "014_sku_classification_and_master_data.sql"

# File paths
DEPLETION_FILE = DOWNLOADS / "ItemActivityInventoryDepletion.xlsx"
VIEW_ITEM_FILE = DOWNLOADS / "ViewItem (3).xlsx"
ACTIVITY_FILES = [DOWNLOADS / f"Item_Activity_Report ({i}).xlsx" for i in range(1, 7)]

# ABC thresholds (cumulative % of annual consumption value)
ABC_A_THRESHOLD = 0.80  # Top 80% = A
ABC_B_THRESHOLD = 0.96  # 80-96% = B, rest = C

# XYZ thresholds (coefficient of variation)
XYZ_X_THRESHOLD = 0.5   # CV < 0.5 = X (stable)
XYZ_Y_THRESHOLD = 1.0   # 0.5 <= CV < 1.0 = Y (moderate), >= 1.0 = Z (erratic)

# Minimum months of data required for XYZ calculation
MIN_MONTHS_FOR_XYZ = 6

# Supplier name mapping (from WMS names to our codes)
SUPPLIER_MAP = {
    'alliance metal changzhou': 'AMC',
    'amc': 'AMC',
    'hx/ whi': 'HX',
    'hx/whi': 'HX',
    'zhongxing': 'ZhongXing',
    'zhong xing': 'ZhongXing',
    'tianjin': 'TianJin',
    'tianjin/whi': 'TianJin',
    'tianijn': 'TianJin',
    'winschem': 'WINSCHEM',
    'changzhou winschem': 'WINSCHEM',
    'changzhou nuode': 'Nuode',
    'nuode': 'Nuode',
}


# ─── Data Structures ─────────────────────────────────────────────────────────

class SKUData:
    """Aggregated data for a single SKU."""
    def __init__(self, sku_code: str):
        self.sku_code = sku_code
        self.supplier = ''
        self.warehouse = ''
        # From ItemActivityInventoryDepletion
        self.avg_ship_wk = 0.0
        self.wks_on_hand = 0.0
        self.beginning_inv = 0
        self.ending_inv = 0
        self.received = 0
        self.shipments = 0
        # From ViewItem
        self.unit_cost = 0.0
        self.weight_lbs = 0.0
        self.length_in = 0.0
        self.width_in = 0.0
        self.height_in = 0.0
        self.dim_uom_units = 1.0  # # Units per Dim UOM
        # Computed
        self.annual_value = 0.0
        self.monthly_qty_out: dict[str, float] = {}  # "2024-01" -> qty
        self.cv = None  # Coefficient of variation
        self.abc_class = ''
        self.xyz_class = ''
        self.xyz_estimated = False  # True if XYZ was estimated (not enough data)


# ─── Parsers ──────────────────────────────────────────────────────────────────

def resolve_supplier(raw_name: str) -> str:
    """Map WMS supplier/section name to our supplier code."""
    lower = raw_name.lower().strip()
    for key, code in SUPPLIER_MAP.items():
        if key in lower:
            return code
    return raw_name.strip()


def parse_depletion(filepath: Path) -> dict[str, SKUData]:
    """Parse ItemActivityInventoryDepletion.xlsx → {sku_code: SKUData}"""
    print(f"\n{'='*60}")
    print(f"Parsing: {filepath.name}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    skus: dict[str, SKUData] = {}
    current_warehouse = ''
    current_supplier = ''

    for row in ws.iter_rows(min_row=1, values_only=True):
        cell0 = str(row[0] or '').strip()

        # Track warehouse sections
        if cell0.startswith('Warehouse:'):
            current_warehouse = cell0.replace('Warehouse:', '').strip()
            continue

        # Skip headers and totals
        if (not cell0 or cell0 == 'SKU' or cell0.startswith('Total') or
            cell0.startswith('Grand') or cell0.startswith('Item Activity')):
            continue

        # Check if this is a supplier name line (no data in Col[2] = Beginning Inventory)
        if row[2] is None:
            # Check if it has "Item Activity From" in col[3] → supplier line
            col3 = str(row[3] or '')
            if 'Item Activity From' in col3 or not any(isinstance(row[i], (int, float)) for i in range(2, 18) if i < len(row)):
                current_supplier = resolve_supplier(cell0)
                continue

        # This should be a SKU data row
        sku_code = cell0
        avg_ship = row[16] if len(row) > 16 and isinstance(row[16], (int, float)) else 0.0
        woh = row[17] if len(row) > 17 and isinstance(row[17], (int, float)) else 0.0
        beg_inv = row[2] if len(row) > 2 and isinstance(row[2], (int, float)) else 0
        received = row[5] if len(row) > 5 and isinstance(row[5], (int, float)) else 0
        shipments = row[9] if len(row) > 9 and isinstance(row[9], (int, float)) else 0
        ending_inv = row[13] if len(row) > 13 and isinstance(row[13], (int, float)) else 0

        # Create or merge SKU data (same SKU can appear in different warehouses)
        if sku_code in skus:
            existing = skus[sku_code]
            existing.avg_ship_wk += avg_ship
            existing.beginning_inv += beg_inv
            existing.ending_inv += ending_inv
            existing.received += received
            existing.shipments += shipments
            # Recalculate WoH as weighted average would be complex; keep the one with more activity
            if avg_ship > existing.avg_ship_wk - avg_ship:
                existing.wks_on_hand = woh
                existing.warehouse = current_warehouse
        else:
            sku = SKUData(sku_code)
            sku.supplier = current_supplier
            sku.warehouse = current_warehouse
            sku.avg_ship_wk = avg_ship
            sku.wks_on_hand = woh
            sku.beginning_inv = int(beg_inv)
            sku.ending_inv = int(ending_inv)
            sku.received = int(received)
            sku.shipments = int(shipments)
            skus[sku_code] = sku

    print(f"  Parsed {len(skus)} unique SKUs")

    # Print summary by supplier
    by_supplier = defaultdict(list)
    for sku in skus.values():
        by_supplier[sku.supplier].append(sku)

    for supplier, supplier_skus in sorted(by_supplier.items()):
        total_ship = sum(s.avg_ship_wk for s in supplier_skus)
        print(f"  {supplier}: {len(supplier_skus)} SKUs, total Avg Ship/Wk = {total_ship:.1f}")

    wb.close()
    return skus


def parse_view_item(filepath: Path) -> dict[str, dict]:
    """Parse ViewItem (3).xlsx → {sku_code: {cost, weight_lbs, length/width/height}}"""
    print(f"\n{'='*60}")
    print(f"Parsing: {filepath.name}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    items: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_sku = None
        current_data: dict = {}
        supplier_name = ''

        skip_prefixes = ('SKU', 'Dimensional', 'Storage', 'Track By', 'View Item')
        next_is_dimensions = False  # Flag: next empty-cell0 row has imperial dims

        for row in ws.iter_rows(min_row=1, values_only=True):
            cell0 = str(row[0] or '').strip()

            # Skip known header/label rows
            if cell0.startswith(skip_prefixes):
                if cell0 == 'Dimensional UOM':
                    next_is_dimensions = True
                else:
                    next_is_dimensions = False
                continue

            # Handle non-empty cell0 (either supplier header or SKU row)
            if cell0:
                next_is_dimensions = False
                cost_val = row[9] if len(row) > 9 else None

                if isinstance(cost_val, (int, float)) and cost_val > 0:
                    # This is a SKU row (has a numeric Cost value)
                    current_sku = cell0
                    items[current_sku] = {
                        'cost': float(cost_val),
                        'description': str(row[1] or ''),
                        'weight_lbs': 0.0,
                        'length_in': 0.0,
                        'width_in': 0.0,
                        'height_in': 0.0,
                        'dim_uom_units': 1.0,
                        'supplier': resolve_supplier(supplier_name),
                    }
                else:
                    # Likely a supplier/section header (e.g., "TianJin/WHI - Kent")
                    supplier_name = cell0
                continue

            # Handle empty cell0 rows (dimensional data, metric data, storage data)
            if not cell0 and current_sku and current_sku in items:
                if next_is_dimensions:
                    # This is the imperial dimensions row
                    # Cols: [2]=Dim UOM, [3]=# Units, [6]=Length, [7]=Width, [8]=Height, [10]=Weight
                    dim_uom = str(row[2] or '').strip() if len(row) > 2 else ''
                    dim_units = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) else None

                    if dim_uom and dim_units:
                        length = row[6] if len(row) > 6 and isinstance(row[6], (int, float)) else 0
                        width = row[7] if len(row) > 7 and isinstance(row[7], (int, float)) else 0
                        height = row[8] if len(row) > 8 and isinstance(row[8], (int, float)) else 0
                        weight = row[10] if len(row) > 10 and isinstance(row[10], (int, float)) else 0

                        if length > 0 and items[current_sku]['length_in'] == 0:
                            items[current_sku]['length_in'] = float(length)
                            items[current_sku]['width_in'] = float(width)
                            items[current_sku]['height_in'] = float(height)
                            items[current_sku]['weight_lbs'] = float(weight)
                            items[current_sku]['dim_uom_units'] = float(dim_units)
                    next_is_dimensions = False

    print(f"  Parsed {len(items)} SKUs across {len(wb.sheetnames)} sheets")

    # Print cost range
    costs = [v['cost'] for v in items.values() if v['cost'] > 0]
    if costs:
        print(f"  Cost range: ${min(costs):.2f} - ${max(costs):.2f}")
        print(f"  SKUs with cost > $100: {sum(1 for c in costs if c > 100)}")

    wb.close()
    return items


def parse_activity_reports(filepaths: list[Path]) -> dict[str, dict[str, float]]:
    """Parse Item Activity Reports → {sku_code: {"2024-01": qty_out, ...}}

    IMPORTANT: Filters out:
      - Inventory adjustments (ref# containing 'adjust'/'adj')
      - Warehouse moves (ref# starting with 'mv')
      - Beginning/Ending Balance rows
    """
    print(f"\n{'='*60}")
    print(f"Parsing: {len(filepaths)} Item Activity Reports")
    print(f"{'='*60}")

    # {sku_code: {month_key: total_qty_out}}
    monthly_out: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    total_transactions = 0
    filtered_adjustments = 0
    filtered_moves = 0

    for fpath in filepaths:
        if not fpath.exists():
            print(f"  SKIP: {fpath.name} not found")
            continue

        print(f"  Processing: {fpath.name}")
        wb = openpyxl.load_workbook(str(fpath), data_only=True)
        ws = wb.active

        current_sku = None
        file_tx = 0
        file_filtered = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            # Detect SKU header rows (col[1] has SKU code, col[4] has qty like '1')
            col1 = str(row[1] or '').strip() if len(row) > 1 else ''

            # SKU identifier row: col[1] is alphanumeric, col[4] is a number (qty indicator)
            if col1 and col1 not in ('', 'Totals:', 'SKU') and len(row) > 4:
                col4 = row[4]
                if isinstance(col4, (int, float)) and col4 > 0:
                    # This is a SKU header row
                    current_sku = col1
                    continue

            if not current_sku:
                continue

            # Parse transaction rows
            date_str = str(row[5] or '').strip() if len(row) > 5 else ''

            # Skip non-date rows
            if not date_str or date_str in ('Activity Date', 'Beginning Balance', 'Ending Balance', ''):
                continue

            # Parse date
            try:
                if isinstance(row[5], datetime):
                    dt = row[5]
                else:
                    dt = datetime.strptime(date_str, "%m/%d/%Y")
            except (ValueError, TypeError):
                continue

            # Get ref# for filtering
            ref = str(row[7] or '').strip().lower() if len(row) > 7 else ''

            # FILTER: Inventory adjustments
            if 'adjust' in ref or 'adj' in ref:
                filtered_adjustments += 1
                file_filtered += 1
                continue

            # FILTER: Warehouse moves
            if ref.startswith('mv'):
                filtered_moves += 1
                file_filtered += 1
                continue

            # Get qty out
            qty_out = row[10] if len(row) > 10 and isinstance(row[10], (int, float)) else 0

            if qty_out > 0:
                month_key = f"{dt.year}-{dt.month:02d}"
                monthly_out[current_sku][month_key] += qty_out
                file_tx += 1
                total_transactions += 1

        print(f"    → {file_tx} outbound transactions, {file_filtered} filtered")
        wb.close()

    print(f"\n  Total: {total_transactions} transactions across {len(monthly_out)} SKUs")
    print(f"  Filtered: {filtered_adjustments} adjustments + {filtered_moves} moves = {filtered_adjustments + filtered_moves}")

    return dict(monthly_out)


# ─── Classification Logic ─────────────────────────────────────────────────────

def compute_abc(skus: dict[str, SKUData], costs: dict[str, dict]):
    """Compute ABC classification based on annual consumption value."""
    print(f"\n{'='*60}")
    print(f"Computing ABC Classification")
    print(f"{'='*60}")

    # Calculate annual value for each SKU
    for sku_code, sku in skus.items():
        # Find cost (try exact match, then without GT suffix)
        cost = 0.0
        if sku_code in costs:
            cost = costs[sku_code]['cost']
        elif sku_code.rstrip('GT') in costs:
            cost = costs[sku_code.rstrip('GT')]['cost']
        elif sku_code + 'GT' in costs:
            cost = costs[sku_code + 'GT']['cost']

        sku.unit_cost = cost
        sku.annual_value = sku.avg_ship_wk * 52 * cost

    # Sort by annual value descending
    sorted_skus = sorted(skus.values(), key=lambda s: s.annual_value, reverse=True)
    total_value = sum(s.annual_value for s in sorted_skus)

    if total_value <= 0:
        print("  WARNING: Total annual value is 0! Check cost data.")
        for sku in sorted_skus:
            sku.abc_class = 'C'
        return

    # Assign ABC based on cumulative percentage
    cumulative = 0.0
    a_count = b_count = c_count = 0

    for sku in sorted_skus:
        cumulative += sku.annual_value
        pct = cumulative / total_value

        if pct <= ABC_A_THRESHOLD:
            sku.abc_class = 'A'
            a_count += 1
        elif pct <= ABC_B_THRESHOLD:
            sku.abc_class = 'B'
            b_count += 1
        else:
            sku.abc_class = 'C'
            c_count += 1

    # Edge case: ensure at least 1 A
    if a_count == 0 and sorted_skus:
        sorted_skus[0].abc_class = 'A'
        a_count = 1
        if sorted_skus[0].abc_class == 'B':
            b_count -= 1
        else:
            c_count -= 1

    print(f"  Total annual consumption value: ${total_value:,.2f}")
    print(f"  A: {a_count} SKUs ({a_count/len(sorted_skus)*100:.0f}%)")
    print(f"  B: {b_count} SKUs ({b_count/len(sorted_skus)*100:.0f}%)")
    print(f"  C: {c_count} SKUs ({c_count/len(sorted_skus)*100:.0f}%)")

    # Print top 10
    print(f"\n  Top 10 SKUs by Annual Value:")
    for i, sku in enumerate(sorted_skus[:10]):
        print(f"    {i+1}. {sku.sku_code} [{sku.abc_class}] "
              f"${sku.annual_value:,.2f}/yr "
              f"(Avg {sku.avg_ship_wk:.1f}/wk × ${sku.unit_cost:.2f})")


def compute_xyz(skus: dict[str, SKUData], monthly_data: dict[str, dict[str, float]]):
    """Compute XYZ classification based on coefficient of variation."""
    print(f"\n{'='*60}")
    print(f"Computing XYZ Classification")
    print(f"{'='*60}")

    x_count = y_count = z_count = 0
    estimated_count = 0

    for sku_code, sku in skus.items():
        # Try to find monthly data
        monthly = monthly_data.get(sku_code, {})

        # Also try without GT suffix
        if not monthly:
            monthly = monthly_data.get(sku_code.rstrip('GT'), {})
        if not monthly:
            monthly = monthly_data.get(sku_code + 'GT', {})

        if len(monthly) >= MIN_MONTHS_FOR_XYZ:
            # Calculate CV from monthly outbound quantities
            values = list(monthly.values())
            mean_val = statistics.mean(values)

            if mean_val > 0:
                std_val = statistics.stdev(values) if len(values) > 1 else 0
                sku.cv = std_val / mean_val
            else:
                sku.cv = float('inf')

            sku.monthly_qty_out = monthly
        else:
            # Estimate: if AvgShip/Wk > 0, assume moderate variability
            # SKUs with very low volume tend to be erratic
            if sku.avg_ship_wk >= 10:
                sku.cv = 0.6  # Assume moderate
            elif sku.avg_ship_wk >= 1:
                sku.cv = 0.8  # Assume moderate-high
            elif sku.avg_ship_wk > 0:
                sku.cv = 1.2  # Assume erratic
            else:
                sku.cv = float('inf')
            sku.xyz_estimated = True
            estimated_count += 1

        # Classify
        if sku.cv is not None:
            if sku.cv < XYZ_X_THRESHOLD:
                sku.xyz_class = 'X'
                x_count += 1
            elif sku.cv < XYZ_Y_THRESHOLD:
                sku.xyz_class = 'Y'
                y_count += 1
            else:
                sku.xyz_class = 'Z'
                z_count += 1

    print(f"  X (stable, CV<{XYZ_X_THRESHOLD}): {x_count} SKUs ({x_count/len(skus)*100:.0f}%)")
    print(f"  Y (moderate, CV<{XYZ_Y_THRESHOLD}): {y_count} SKUs ({y_count/len(skus)*100:.0f}%)")
    print(f"  Z (erratic, CV>={XYZ_Y_THRESHOLD}): {z_count} SKUs ({z_count/len(skus)*100:.0f}%)")
    print(f"  Estimated (insufficient monthly data): {estimated_count} SKUs")

    # Print examples
    print(f"\n  Sample CV values:")
    sorted_by_cv = sorted([s for s in skus.values() if s.cv is not None and s.cv != float('inf')],
                          key=lambda s: s.cv)
    for sku in sorted_by_cv[:5]:
        est = " (estimated)" if sku.xyz_estimated else ""
        months = len(sku.monthly_qty_out)
        print(f"    {sku.sku_code} [{sku.xyz_class}] CV={sku.cv:.3f} ({months} months){est}")
    if len(sorted_by_cv) > 5:
        print(f"    ...")
        for sku in sorted_by_cv[-3:]:
            est = " (estimated)" if sku.xyz_estimated else ""
            months = len(sku.monthly_qty_out)
            print(f"    {sku.sku_code} [{sku.xyz_class}] CV={sku.cv:.3f} ({months} months){est}")


def merge_master_data(skus: dict[str, SKUData], view_items: dict[str, dict]):
    """Merge ViewItem data (cost, weight, dimensions) into SKU records."""
    matched = 0
    unmatched_skus = []

    for sku_code, sku in skus.items():
        item = None
        # Try exact match
        if sku_code in view_items:
            item = view_items[sku_code]
        # Try without GT suffix
        elif sku_code.rstrip('GT') in view_items:
            item = view_items[sku_code.rstrip('GT')]
        # Try with GT suffix
        elif sku_code + 'GT' in view_items:
            item = view_items[sku_code + 'GT']

        if item:
            sku.unit_cost = item['cost']
            sku.weight_lbs = item['weight_lbs']
            sku.length_in = item['length_in']
            sku.width_in = item['width_in']
            sku.height_in = item['height_in']
            sku.dim_uom_units = item.get('dim_uom_units', 1.0)
            if not sku.supplier:
                sku.supplier = item.get('supplier', '')
            matched += 1
        else:
            unmatched_skus.append(sku_code)

    print(f"\n  Master data merge: {matched}/{len(skus)} matched")
    if unmatched_skus:
        print(f"  Unmatched SKUs: {unmatched_skus[:20]}")


# ─── SQL Generation ───────────────────────────────────────────────────────────

def generate_sql(skus: dict[str, SKUData], output_path: Path):
    """Generate SQL migration file with classification results."""
    print(f"\n{'='*60}")
    print(f"Generating SQL: {output_path.name}")
    print(f"{'='*60}")

    lines = []
    lines.append("-- " + "=" * 75)
    lines.append("-- 014: Phase 3A — SKU Classification & Master Data Import")
    lines.append("--")
    lines.append("-- Auto-generated by classify_skus.py")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("--")
    lines.append("-- Data sources:")
    lines.append("--   ItemActivityInventoryDepletion.xlsx (Avg Ship/Wk)")
    lines.append("--   ViewItem (3).xlsx (Cost, Weight, Dimensions)")
    lines.append("--   Item_Activity_Report (1-6).xlsx (Monthly CV)")
    lines.append("--")
    lines.append("-- Prerequisites: 001-013 must have been run")
    lines.append("-- " + "=" * 75)
    lines.append("")

    # ── Part 1: Add new columns ──
    lines.append("")
    lines.append("-- " + "=" * 45)
    lines.append("-- PART 1: Add new columns to skus table")
    lines.append("-- " + "=" * 45)
    lines.append("")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS unit_cost NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS annual_consumption_value NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS avg_weekly_demand NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS cv_demand NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS dimensions_cbm NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS length_in NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS width_in NUMERIC;")
    lines.append("ALTER TABLE skus ADD COLUMN IF NOT EXISTS height_in NUMERIC;")
    lines.append("")

    # ── Part 2: Upsert SKU data ──
    lines.append("")
    lines.append("-- " + "=" * 45)
    lines.append("-- PART 2: Update/Insert SKU master data + classification")
    lines.append("-- " + "=" * 45)
    lines.append("")

    sorted_skus = sorted(skus.values(), key=lambda s: s.annual_value, reverse=True)

    for sku in sorted_skus:
        if not sku.abc_class:
            continue

        # Calculate CBM
        cbm = 0.0
        if sku.length_in > 0 and sku.width_in > 0 and sku.height_in > 0:
            # Convert inches to meters and calculate volume
            cbm = (sku.length_in * 0.0254) * (sku.width_in * 0.0254) * (sku.height_in * 0.0254)

        cv_str = f"{sku.cv:.4f}" if sku.cv is not None and sku.cv != float('inf') else "NULL"

        # Use sku_code for matching (try exact, might need GT suffix handling)
        sku_code_sql = sku.sku_code.replace("'", "''")

        lines.append(f"-- {sku.sku_code}: {sku.abc_class}{sku.xyz_class} | "
                      f"${sku.annual_value:,.0f}/yr | "
                      f"Avg {sku.avg_ship_wk:.1f}/wk | "
                      f"CV={cv_str} | "
                      f"Cost ${sku.unit_cost:.2f}"
                      + (" (XYZ estimated)" if sku.xyz_estimated else ""))

        # Try to UPDATE existing SKU first, handle both with and without GT
        lines.append(f"UPDATE skus SET")
        lines.append(f"  unit_cost = {sku.unit_cost:.4f},")
        lines.append(f"  unit_weight = {sku.weight_lbs:.4f},")
        lines.append(f"  abc_class = '{sku.abc_class}',")
        lines.append(f"  xyz_class = '{sku.xyz_class}',")
        lines.append(f"  annual_consumption_value = {sku.annual_value:.2f},")
        lines.append(f"  avg_weekly_demand = {sku.avg_ship_wk:.4f},")
        lines.append(f"  cv_demand = {cv_str},")
        if cbm > 0:
            lines.append(f"  dimensions_cbm = {cbm:.6f},")
            lines.append(f"  length_in = {sku.length_in:.2f},")
            lines.append(f"  width_in = {sku.width_in:.2f},")
            lines.append(f"  height_in = {sku.height_in:.2f},")
        if sku.supplier:
            lines.append(f"  supplier_code = '{sku.supplier}',")
        lines.append(f"  updated_at = NOW()")
        lines.append(f"WHERE sku_code = '{sku_code_sql}'")
        lines.append(f"   OR sku_code = '{sku_code_sql.rstrip('GT')}'")
        lines.append(f"   OR sku_code = '{sku_code_sql}GT';")
        lines.append("")

    # ── Part 3: Classification view ──
    lines.append("")
    lines.append("-- " + "=" * 45)
    lines.append("-- PART 3: SKU Classification View")
    lines.append("-- " + "=" * 45)
    lines.append("")
    lines.append("""CREATE OR REPLACE VIEW v_sku_classification AS
SELECT
  s.id,
  s.sku_code,
  s.part_model,
  s.description,
  s.supplier_code,
  s.abc_class,
  s.xyz_class,
  COALESCE(s.abc_class, '') || COALESCE(s.xyz_class, '') as matrix_cell,
  s.unit_cost,
  s.annual_consumption_value,
  s.avg_weekly_demand,
  s.cv_demand,
  s.safety_stock_weeks,
  s.reorder_point,
  s.moq,
  s.lead_time_weeks,
  s.unit_weight,
  s.qty_per_container,
  s.dimensions_cbm,
  s.length_in,
  s.width_in,
  s.height_in
FROM skus s
WHERE s.abc_class IS NOT NULL
ORDER BY
  CASE s.abc_class WHEN 'A' THEN 1 WHEN 'B' THEN 2 ELSE 3 END,
  CASE s.xyz_class WHEN 'X' THEN 1 WHEN 'Y' THEN 2 ELSE 3 END,
  s.annual_consumption_value DESC NULLS LAST;""")
    lines.append("")

    # ── Part 4: Classification policies table ──
    lines.append("")
    lines.append("-- " + "=" * 45)
    lines.append("-- PART 4: Classification Policies Table (9-Grid)")
    lines.append("-- " + "=" * 45)
    lines.append("")
    lines.append("""CREATE TABLE IF NOT EXISTS classification_policies (
  id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
  matrix_cell TEXT NOT NULL UNIQUE,
  service_level NUMERIC NOT NULL,
  target_woh NUMERIC NOT NULL,
  review_frequency TEXT NOT NULL,
  replenishment_method TEXT NOT NULL,
  safety_stock_multiplier NUMERIC DEFAULT 1.0,
  notes TEXT,
  created_at TIMESTAMPTZ DEFAULT NOW(),
  updated_at TIMESTAMPTZ DEFAULT NOW()
);""")
    lines.append("")
    lines.append("""-- Netstock-recommended default policies
INSERT INTO classification_policies (matrix_cell, service_level, target_woh, review_frequency, replenishment_method, notes) VALUES
  ('AX', 0.97, 4,  'weekly',    'auto',          'High value + stable: tight control, auto replenish'),
  ('AY', 0.95, 5,  'weekly',    'auto',          'High value + moderate: buffer slightly more'),
  ('AZ', 0.93, 6,  'weekly',    'manual_review', 'High value + erratic: human review before ordering'),
  ('BX', 0.95, 5,  'biweekly',  'auto',          'Medium value + stable: standard auto'),
  ('BY', 0.93, 6,  'biweekly',  'auto',          'Medium value + moderate: moderate buffer'),
  ('BZ', 0.90, 8,  'biweekly',  'manual_review', 'Medium value + erratic: review before ordering'),
  ('CX', 0.92, 6,  'monthly',   'auto',          'Low value + stable: less frequent review'),
  ('CY', 0.90, 8,  'monthly',   'auto',          'Low value + moderate: bulk order'),
  ('CZ', 0.85, 10, 'monthly',   'on_demand',     'Low value + erratic: order only when needed')
ON CONFLICT (matrix_cell) DO UPDATE SET
  service_level = EXCLUDED.service_level,
  target_woh = EXCLUDED.target_woh,
  review_frequency = EXCLUDED.review_frequency,
  replenishment_method = EXCLUDED.replenishment_method,
  notes = EXCLUDED.notes,
  updated_at = NOW();""")
    lines.append("")

    # ── Part 5: Permissions ──
    lines.append("")
    lines.append("-- " + "=" * 45)
    lines.append("-- PART 5: Permissions")
    lines.append("-- " + "=" * 45)
    lines.append("")
    lines.append("GRANT ALL ON classification_policies TO authenticated;")
    lines.append("GRANT SELECT ON v_sku_classification TO authenticated;")
    lines.append("")

    # Write file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"  Written {len(lines)} lines to {output_path.name}")


# ─── Summary ──────────────────────────────────────────────────────────────────

def print_matrix_summary(skus: dict[str, SKUData]):
    """Print 9-grid matrix summary."""
    print(f"\n{'='*60}")
    print(f"ABC/XYZ 9-Grid Matrix Summary")
    print(f"{'='*60}")

    matrix = defaultdict(list)
    for sku in skus.values():
        if sku.abc_class and sku.xyz_class:
            cell = f"{sku.abc_class}{sku.xyz_class}"
            matrix[cell].append(sku)

    print(f"\n         {'X (Stable)':^20} {'Y (Moderate)':^20} {'Z (Erratic)':^20}")
    print(f"    {'─'*62}")

    for abc in ['A', 'B', 'C']:
        row_parts = []
        for xyz in ['X', 'Y', 'Z']:
            cell = f"{abc}{xyz}"
            cell_skus = matrix.get(cell, [])
            total_val = sum(s.annual_value for s in cell_skus)
            row_parts.append(f"{len(cell_skus):>3} SKUs ${total_val:>10,.0f}")

        label = {'A': 'A (High)', 'B': 'B (Med)', 'C': 'C (Low)'}[abc]
        print(f" {label:>8} │ {row_parts[0]:^20}│ {row_parts[1]:^20}│ {row_parts[2]:^20}│")
        if abc != 'C':
            print(f"         {'─'*62}")

    print(f"    {'─'*62}")

    # Print SKU list per cell
    print(f"\n  Detailed breakdown:")
    for abc in ['A', 'B', 'C']:
        for xyz in ['X', 'Y', 'Z']:
            cell = f"{abc}{xyz}"
            cell_skus = sorted(matrix.get(cell, []), key=lambda s: s.annual_value, reverse=True)
            if cell_skus:
                print(f"\n  [{cell}] {len(cell_skus)} SKUs:")
                for sku in cell_skus:
                    cv_str = f"CV={sku.cv:.2f}" if sku.cv and sku.cv != float('inf') else "CV=N/A"
                    est = "*" if sku.xyz_estimated else ""
                    print(f"    {sku.sku_code:<15} {sku.supplier:<12} "
                          f"${sku.annual_value:>10,.0f}/yr  "
                          f"Avg {sku.avg_ship_wk:>7.1f}/wk  "
                          f"{cv_str}{est}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Phase 3A: ABC/XYZ SKU Classification Engine")
    print("=" * 60)

    # Validate input files
    missing = []
    if not DEPLETION_FILE.exists():
        missing.append(str(DEPLETION_FILE))
    if not VIEW_ITEM_FILE.exists():
        missing.append(str(VIEW_ITEM_FILE))
    for f in ACTIVITY_FILES:
        if not f.exists():
            missing.append(str(f))

    if missing:
        print(f"\nWARNING: Missing files:")
        for f in missing:
            print(f"  - {f}")
        if not DEPLETION_FILE.exists() or not VIEW_ITEM_FILE.exists():
            print("\nCritical files missing. Cannot proceed.")
            sys.exit(1)
        print("\nProceeding with available files...\n")

    # Step 1: Parse all data sources
    skus = parse_depletion(DEPLETION_FILE)
    view_items = parse_view_item(VIEW_ITEM_FILE)
    monthly_data = parse_activity_reports([f for f in ACTIVITY_FILES if f.exists()])

    # Step 2: Merge master data
    merge_master_data(skus, view_items)

    # Step 3: Compute ABC
    compute_abc(skus, view_items)

    # Step 4: Compute XYZ
    compute_xyz(skus, monthly_data)

    # Step 5: Print summary
    print_matrix_summary(skus)

    # Step 6: Generate SQL
    generate_sql(skus, OUTPUT_SQL)

    print(f"\n{'='*60}")
    print(f"DONE!")
    print(f"{'='*60}")
    print(f"\nNext steps:")
    print(f"  1. Review generated SQL: {OUTPUT_SQL}")
    print(f"  2. Push to GitHub")
    print(f"  3. Chris executes migration 014 on Supabase")
    print(f"  4. Verify: SELECT * FROM v_sku_classification;")


if __name__ == '__main__':
    main()
